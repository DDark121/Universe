from __future__ import annotations

import asyncio
import csv
import hashlib
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.time import utc_now
from app.db.enums import (
    AttendanceStatus,
    ExportFormat,
    ExportJobType,
    JobStatus,
    LessonStatus,
    RoleCode,
)
from app.db.models import (
    AttendanceRecord,
    Discipline,
    ExportJob,
    Group,
    ImportJob,
    Lesson,
    RiskCard,
    Role,
    StudentGroupMembership,
    User,
)
from app.services.import_apply import (
    ensure_student_membership,
    ensure_teacher_assignment,
    resolve_discipline,
    resolve_group,
    resolve_user,
    upsert_lesson,
)

EXPORTS_DIR = Path("/tmp/universe-exports")
IMPORTS_DIR = Path("/tmp/universe-imports")
IMPORT_ERRORS_DIR = Path("/tmp/universe-import-errors")


def _ensure_dirs() -> None:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    IMPORTS_DIR.mkdir(parents=True, exist_ok=True)
    IMPORT_ERRORS_DIR.mkdir(parents=True, exist_ok=True)


def _normalize_key(raw: str) -> str:
    return raw.strip().lower().replace(" ", "_")


def _first_value(row: dict[str, Any], keys: set[str]) -> str | None:
    for key in keys:
        if key in row and row[key] is not None:
            value = str(row[key]).strip()
            if value:
                return value
    return None


def _parse_datetime(value: str) -> datetime:
    value = value.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M"):
        try:
            parsed = datetime.strptime(value, fmt)
            return parsed.replace(tzinfo=UTC)
        except ValueError:
            continue
    parsed = datetime.fromisoformat(value)
    return parsed if parsed.tzinfo is not None else parsed.replace(tzinfo=UTC)


def _generated_code(prefix: str, value: str) -> str:
    normalized = re.sub(r"[^0-9A-Za-zА-Яа-я_.-]+", "_", value.strip()).strip("_")
    if normalized:
        return normalized.upper()[:64]
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:10]
    return f"{prefix}-{digest}".upper()


def _generated_username(full_name: str) -> str:
    digest = hashlib.sha1(full_name.encode("utf-8")).hexdigest()[:10]
    return f"teacher_{digest}"


def _parse_schedule_datetimes(row: dict[str, Any]) -> tuple[datetime, datetime]:
    starts_at_raw = _first_value(row, {"starts_at", "начало"})
    ends_at_raw = _first_value(row, {"ends_at", "конец"})
    if starts_at_raw and ends_at_raw:
        return _parse_datetime(starts_at_raw), _parse_datetime(ends_at_raw)

    date_raw = _first_value(row, {"date", "lesson_date", "дата"})
    start_time_raw = _first_value(row, {"start_time", "time_start", "время_начала", "начало_пары"})
    end_time_raw = _first_value(row, {"end_time", "time_end", "время_конца", "конец_пары"})
    if not date_raw or not start_time_raw or not end_time_raw:
        raise ValueError("starts_at/ends_at or date/start_time/end_time are required")

    date_part = date_raw.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            parsed_date = datetime.strptime(date_part, fmt).date()
            break
        except ValueError:
            parsed_date = None
    if parsed_date is None:
        parsed_date = datetime.fromisoformat(date_part).date()

    return (
        _parse_datetime(f"{parsed_date.isoformat()} {start_time_raw.strip()}"),
        _parse_datetime(f"{parsed_date.isoformat()} {end_time_raw.strip()}"),
    )


def _load_rows(file_path: str) -> list[dict[str, str]]:
    path = Path(file_path)
    ext = path.suffix.lower()
    if ext == ".csv":
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = []
            for row in reader:
                normalized = {_normalize_key(key): (value or "") for key, value in row.items() if key}
                rows.append(normalized)
            return rows

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return []
    headers = [_normalize_key(str(item or "")) for item in all_rows[0]]
    rows = []
    for values in all_rows[1:]:
        record: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = str(values[idx] or "")
        rows.append(record)
    return rows


def _write_error_report(job_id: UUID, rows: list[dict[str, Any]]) -> str:
    _ensure_dirs()
    path = IMPORT_ERRORS_DIR / f"import_errors_{job_id}.csv"
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["row", "error"])
        writer.writeheader()
        writer.writerows(rows)
    return str(path)


async def load_rows_async(file_path: str) -> list[dict[str, str]]:
    return await asyncio.to_thread(_load_rows, file_path)


async def write_error_report_async(job_id: UUID, rows: list[dict[str, Any]]) -> str:
    return await asyncio.to_thread(_write_error_report, job_id, rows)


def _parse_role_codes(raw_roles: str) -> list[RoleCode]:
    mapping = {
        "student": RoleCode.STUDENT,
        "студент": RoleCode.STUDENT,
        "teacher": RoleCode.TEACHER,
        "преподаватель": RoleCode.TEACHER,
        "admin": RoleCode.ADMIN,
        "админ": RoleCode.ADMIN,
        "administrator": RoleCode.ADMIN,
        "curator": RoleCode.CURATOR,
        "тьютор": RoleCode.CURATOR,
        "куратор": RoleCode.CURATOR,
    }
    parsed: list[RoleCode] = []
    for item in raw_roles.replace(";", ",").split(","):
        token = item.strip().lower()
        if not token:
            continue
        if token not in mapping:
            raise ValueError(f"Unknown role '{item.strip()}'")
        parsed.append(mapping[token])
    if not parsed:
        raise ValueError("Role list is empty")
    return parsed


async def _roles_by_codes(session: AsyncSession, codes: list[RoleCode]) -> list[Role]:
    rows = (await session.execute(select(Role).where(Role.code.in_(codes)))).scalars().all()
    if len(rows) != len(set(codes)):
        raise ValueError("Some role codes are not configured")
    return rows


@dataclass
class ImportProcessResult:
    total_rows: int
    processed_rows: int
    errors: list[dict[str, Any]]
    error_file_path: str | None


async def process_import_users(session: AsyncSession, job: ImportJob) -> ImportProcessResult:
    rows = await load_rows_async(job.file_path)
    total_rows = len(rows)
    processed_rows = 0
    errors: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=2):
        try:
            username = _first_value(row, {"username", "логин"})
            full_name = _first_value(row, {"full_name", "фио", "full_name_"})
            email = _first_value(row, {"email", "почта", "e-mail"})
            phone_number = _first_value(row, {"phone_number", "phone", "телефон"})
            roles_raw = _first_value(row, {"roles", "role", "роли", "роль"})
            group_code = _first_value(row, {"group_code", "group", "код_группы"})

            if not username or not full_name or not phone_number or not roles_raw:
                raise ValueError("username/full_name/phone_number/roles are required")

            role_codes = _parse_role_codes(roles_raw)
            await _roles_by_codes(session, role_codes)
            user = await resolve_user(
                session,
                username=username,
                full_name=full_name,
                email=email,
                role_codes=role_codes,
                phone_number=phone_number,
                action="upsert",
                role_update_strategy="replace",
            )

            if group_code and RoleCode.STUDENT in role_codes:
                group = (await session.execute(select(Group).where(Group.code == group_code))).scalar_one_or_none()
                if not group:
                    raise ValueError(f"Unknown group_code '{group_code}'")
                await ensure_student_membership(
                    session,
                    student_id=user.id,
                    group_id=group.id,
                    start_date=utc_now().date(),
                )

            processed_rows += 1
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": index, "error": str(exc)})

        job.processed_rows = processed_rows
        job.total_rows = total_rows
        await session.commit()

    error_file_path = await write_error_report_async(job.id, errors) if errors else None
    return ImportProcessResult(
        total_rows=total_rows,
        processed_rows=processed_rows,
        errors=errors,
        error_file_path=error_file_path,
    )


async def process_import_schedule(session: AsyncSession, job: ImportJob) -> ImportProcessResult:
    rows = await load_rows_async(job.file_path)
    total_rows = len(rows)
    processed_rows = 0
    errors: list[dict[str, Any]] = []

    status_map = {
        "planned": LessonStatus.PLANNED,
        "запланировано": LessonStatus.PLANNED,
        "in_progress": LessonStatus.IN_PROGRESS,
        "идет": LessonStatus.IN_PROGRESS,
        "completed": LessonStatus.COMPLETED,
        "завершено": LessonStatus.COMPLETED,
        "canceled": LessonStatus.CANCELED,
        "отменено": LessonStatus.CANCELED,
        "rescheduled": LessonStatus.RESCHEDULED,
        "перенесено": LessonStatus.RESCHEDULED,
    }

    for index, row in enumerate(rows, start=2):
        try:
            group_code = _first_value(row, {"group_code", "код_группы", "group", "группа"})
            group_name = _first_value(row, {"group_name", "название_группы"})
            discipline_code = _first_value(row, {"discipline_code", "код_дисциплины"})
            discipline_name = _first_value(
                row,
                {"discipline_name", "discipline", "дисциплина", "предмет", "название_дисциплины"},
            )
            teacher_username = _first_value(row, {"teacher_username", "логин_преподавателя"})
            teacher_full_name = _first_value(row, {"teacher_name", "teacher", "преподаватель", "фио_преподавателя"})
            teacher_email = _first_value(row, {"teacher_email", "email_преподавателя", "почта_преподавателя"})
            teacher_phone = _first_value(row, {"teacher_phone", "phone_преподавателя", "телефон_преподавателя"})
            room = _first_value(row, {"room", "аудитория"})
            status_raw = _first_value(row, {"status", "статус"}) or "planned"

            if not group_code and not group_name:
                raise ValueError("group_code or group_name is required")
            if not discipline_code and not discipline_name:
                raise ValueError("discipline_code or discipline_name is required")
            if not teacher_username and not teacher_full_name:
                raise ValueError("teacher_username or teacher_name is required")

            group_code = group_code or _generated_code("GROUP", group_name or "")
            group = await resolve_group(
                session,
                code=group_code,
                name=group_name or group_code,
                action="upsert",
            )

            discipline_code = discipline_code or _generated_code("DISC", discipline_name or "")
            discipline = await resolve_discipline(
                session,
                code=discipline_code,
                name=discipline_name or discipline_code,
                action="upsert",
            )

            teacher_username = teacher_username or _generated_username(teacher_full_name or "")
            teacher = await resolve_user(
                session,
                username=teacher_username,
                full_name=teacher_full_name or teacher_username,
                email=teacher_email,
                role_codes=[RoleCode.TEACHER],
                phone_number=teacher_phone,
                action="upsert",
                role_update_strategy="merge",
            )
            if RoleCode.TEACHER not in {role.code for role in teacher.roles}:
                raise ValueError("Referenced user has no teacher role")

            starts_at, ends_at = _parse_schedule_datetimes(row)
            if ends_at <= starts_at:
                raise ValueError("ends_at must be greater than starts_at")

            lesson_status = status_map.get(status_raw.strip().lower())
            if lesson_status is None:
                raise ValueError(f"Unsupported lesson status '{status_raw}'")

            await upsert_lesson(
                session,
                group=group,
                discipline=discipline,
                teacher=teacher,
                starts_at=starts_at,
                ends_at=ends_at,
                room=room,
                status=lesson_status,
            )
            await ensure_teacher_assignment(
                session,
                teacher_id=teacher.id,
                discipline_id=discipline.id,
                group_id=group.id,
            )

            processed_rows += 1
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": index, "error": str(exc)})

        job.processed_rows = processed_rows
        job.total_rows = total_rows
        await session.commit()

    error_file_path = await write_error_report_async(job.id, errors) if errors else None
    return ImportProcessResult(
        total_rows=total_rows,
        processed_rows=processed_rows,
        errors=errors,
        error_file_path=error_file_path,
    )


async def mark_job_processing(session: AsyncSession, job: ExportJob) -> None:
    job.status = JobStatus.PROCESSING
    await session.commit()


async def mark_job_done(session: AsyncSession, job: ExportJob, file_path: str) -> None:
    job.status = JobStatus.DONE
    job.file_path = file_path
    job.completed_at = utc_now()
    await session.commit()


async def mark_job_failed(session: AsyncSession, job: ExportJob) -> None:
    job.status = JobStatus.FAILED
    job.completed_at = utc_now()
    await session.commit()


async def mark_import_job_processing(session: AsyncSession, job: ImportJob) -> None:
    job.status = JobStatus.PROCESSING
    await session.commit()


async def mark_import_job_done(session: AsyncSession, job: ImportJob, result: ImportProcessResult) -> None:
    job.status = JobStatus.DONE if not result.errors else JobStatus.FAILED
    job.total_rows = result.total_rows
    job.processed_rows = result.processed_rows
    job.completed_at = utc_now()
    error_payload: dict[str, Any] | None = None
    if result.errors:
        error_payload = {
            "count": len(result.errors),
            "sample": result.errors[:100],
            "file_path": result.error_file_path,
        }
    job.error_report = error_payload
    await session.commit()


async def mark_import_job_failed(session: AsyncSession, job: ImportJob, detail: str) -> None:
    job.status = JobStatus.FAILED
    job.completed_at = utc_now()
    job.error_report = {"error": detail}
    await session.commit()


def export_rows(file_path: Path, fmt: ExportFormat, rows: list[dict]) -> None:
    if fmt == ExportFormat.CSV:
        with file_path.open("w", newline="", encoding="utf-8") as csvfile:
            if not rows:
                csvfile.write("")
                return
            writer = csv.DictWriter(csvfile, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    wb.save(file_path)


async def export_rows_async(file_path: Path, fmt: ExportFormat, rows: list[dict]) -> None:
    await asyncio.to_thread(export_rows, file_path, fmt, rows)


def build_export_path(export_id: UUID, fmt: ExportFormat) -> Path:
    _ensure_dirs()
    ext = "csv" if fmt == ExportFormat.CSV else "xlsx"
    return EXPORTS_DIR / f"export_{export_id}.{ext}"


async def build_export_path_async(export_id: UUID, fmt: ExportFormat) -> Path:
    return await asyncio.to_thread(build_export_path, export_id, fmt)


async def build_export_rows(session: AsyncSession, job: ExportJob) -> list[dict[str, Any]]:
    filters = job.filters or {}
    if job.job_type == ExportJobType.SCHEDULE:
        stmt = (
            select(Lesson, Group, Discipline, User)
            .join(Group, Group.id == Lesson.group_id)
            .join(Discipline, Discipline.id == Lesson.discipline_id)
            .join(User, User.id == Lesson.teacher_id)
        )
        date_from = filters.get("date_from")
        date_to = filters.get("date_to")
        if date_from:
            stmt = stmt.where(Lesson.starts_at >= _parse_datetime(str(date_from)))
        if date_to:
            stmt = stmt.where(Lesson.starts_at <= _parse_datetime(str(date_to)))
        if filters.get("group_id"):
            stmt = stmt.where(Lesson.group_id == UUID(str(filters["group_id"])))
        if filters.get("group_code"):
            stmt = stmt.where(Group.code == str(filters["group_code"]))
        if filters.get("discipline_id"):
            stmt = stmt.where(Lesson.discipline_id == UUID(str(filters["discipline_id"])))
        if filters.get("teacher_id"):
            stmt = stmt.where(Lesson.teacher_id == UUID(str(filters["teacher_id"])))
        if filters.get("status"):
            stmt = stmt.where(Lesson.status == LessonStatus(str(filters["status"])))

        rows = (await session.execute(stmt.order_by(Lesson.starts_at.asc()))).all()
        return [
            {
                "lesson_id": str(lesson.id),
                "group_id": str(group.id),
                "group_code": group.code,
                "group_name": group.name,
                "discipline_id": str(discipline.id),
                "discipline_code": discipline.code,
                "discipline_name": discipline.name,
                "teacher_id": str(teacher.id),
                "teacher_username": teacher.username,
                "teacher_name": teacher.full_name,
                "starts_at": lesson.starts_at.isoformat(),
                "ends_at": lesson.ends_at.isoformat(),
                "date": lesson.starts_at.date().isoformat(),
                "start_time": lesson.starts_at.strftime("%H:%M"),
                "end_time": lesson.ends_at.strftime("%H:%M"),
                "room": lesson.room,
                "status": lesson.status.value,
            }
            for lesson, group, discipline, teacher in rows
        ]

    if job.job_type == ExportJobType.RISK_LIST:
        stmt = (
            select(RiskCard, User)
            .join(User, User.id == RiskCard.student_id)
            .where(RiskCard.is_active.is_(True))
            .order_by(RiskCard.updated_at.desc())
        )
        if filters.get("group_id"):
            group_id = UUID(str(filters["group_id"]))
            student_ids = (
                await session.execute(
                    select(StudentGroupMembership.student_id).where(
                        StudentGroupMembership.group_id == group_id,
                        StudentGroupMembership.end_date.is_(None),
                    )
                )
            ).all()
            ids = [row[0] for row in student_ids]
            if ids:
                stmt = stmt.where(User.id.in_(ids))
            else:
                return []

        rows = (await session.execute(stmt)).all()
        return [
            {
                "student_id": str(user.id),
                "student_name": user.full_name,
                "score": float(card.last_score),
                "late_count": card.late_count,
                "unexcused_absence_count": card.unexcused_absence_count,
                "reasons": card.reasons,
                "updated_at": card.updated_at.isoformat() if card.updated_at else None,
            }
            for card, user in rows
        ]

    report_type = str(filters.get("report", "attendance")).lower()
    stmt = (
        select(AttendanceRecord, Lesson, User)
        .join(Lesson, Lesson.id == AttendanceRecord.lesson_id)
        .join(User, User.id == AttendanceRecord.student_id)
    )

    date_from = filters.get("date_from")
    date_to = filters.get("date_to")
    if date_from:
        stmt = stmt.where(Lesson.starts_at >= _parse_datetime(str(date_from)))
    if date_to:
        stmt = stmt.where(Lesson.starts_at <= _parse_datetime(str(date_to)))
    if filters.get("student_id"):
        stmt = stmt.where(AttendanceRecord.student_id == UUID(str(filters["student_id"])))
    if filters.get("group_id"):
        stmt = stmt.where(Lesson.group_id == UUID(str(filters["group_id"])))
    if filters.get("discipline_id"):
        stmt = stmt.where(Lesson.discipline_id == UUID(str(filters["discipline_id"])))
    if filters.get("teacher_id"):
        stmt = stmt.where(Lesson.teacher_id == UUID(str(filters["teacher_id"])))

    if report_type == "lates":
        stmt = stmt.where(AttendanceRecord.status == AttendanceStatus.LATE)
    elif report_type == "absences":
        stmt = stmt.where(AttendanceRecord.status == AttendanceStatus.ABSENT)
        if "excused" in filters:
            excused = str(filters.get("excused")).lower() in {"1", "true", "yes"}
            stmt = stmt.where(AttendanceRecord.is_excused == excused)

    rows = (await session.execute(stmt.order_by(Lesson.starts_at.desc()))).all()
    return [
        {
            "attendance_id": str(attendance.id),
            "lesson_id": str(lesson.id),
            "student_id": str(user.id),
            "student_name": user.full_name,
            "group_id": str(lesson.group_id),
            "discipline_id": str(lesson.discipline_id),
            "teacher_id": str(lesson.teacher_id),
            "lesson_starts_at": lesson.starts_at.isoformat(),
            "lesson_ends_at": lesson.ends_at.isoformat(),
            "status": attendance.status.value,
            "source": attendance.source.value,
            "marked_at": attendance.marked_at.isoformat(),
            "is_excused": attendance.is_excused,
            "excused_category": attendance.excused_category,
        }
        for attendance, lesson, user in rows
    ]
